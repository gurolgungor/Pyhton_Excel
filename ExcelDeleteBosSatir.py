#-------------------------------------------------------------------------------
# Name:        ExcelDelete
# Purpose:
#
# Author:      Gürol Güngör
#
# Created:     25.09.2022
# Copyright:   (c) Gürol Güngör 2022
# Licence:     GNU - General Public Licence
#-------------------------------------------------------------------------------

#Excel kitapligini yukluyoruz
import openpyxl
import sys

# bir satirdaki bilgiyi silen fonksiyon
def remove(sheet, row):
	# iterate the row object
	for cell in row:
		# check the value of each cell in
		# the row, if any of the value is not
		# None return without removing the row
		if cell.value != None:
			return
	# get the row number from the first cell
	# and remove the row
	sheet.delete_rows(row[0].row, 1)


def main():
    pass

if __name__ == '__main__':
    main()

#Excel dosyası yeri ve adi tanımlanır
ExcelDosyasi = r"C:\Database\rehber.xlsx"
ExcelSayfasi = r"Sheet1"

# workbook ve worksheet olusturyoruz
workbook  = openpyxl.load_workbook(ExcelDosyasi)
worksheet = workbook[ExcelSayfasi]

print("Boş satırlar silinecek")
print("Silinme işlemi öncesi kayıt sayısı :", worksheet.max_row)

# iterate the sheet object
for row in worksheet:
	remove(worksheet,row)

print("Silinme işlemi sonrası kayıt sayısı :",worksheet.max_row)

# Excel dosyası kayıt ediliyor
workbook.save(ExcelDosyasi)

# Excel dosyası kapatiliyor
workbook.close

# Programdan cikiliyor
sys.exit
