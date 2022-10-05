#-------------------------------------------------------------------------------
# Name:        ExcelCreate
# Purpose:
#
# Author:      Gürol Güngör
#
# Created:     25.09.2022
# Copyright:   (c) Gürol Güngör 2022
# Licence:     GNU - General Public Licence
#-------------------------------------------------------------------------------

#Excel kitapligini yukluyoruz
import openpyxl
import sys

#Excel dosyası yeri ve adi tanımlanır
ExcelDosyasi   = r"C:\Database\rehber.xlsx"
ExcelSayfaismi = r"Sheet1"

# workbook ve worksheet olusturyoruz
workbook = openpyxl.load_workbook(ExcelDosyasi)
worksheet = workbook.get_sheet_by_name(ExcelSayfaismi)

# Excel kayitlarini döngü içerisinde okuyoruz.
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        print(col[i].value, end="\t\t")
    print('')

# workbook cikmadan once kapatiyoruz.
workbook.close()

# programi kapatiyoruz.
sys.exit