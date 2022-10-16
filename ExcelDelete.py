#-------------------------------------------------------------------------------
# Name:        ExcelDelete
# Purpose:
#
# Author:      Gürol Güngör
#
# Created:     25.09.2022
# Copyright:   (c) Gürol Güngör 2022
# Licence:     GNU - General Public Licence
#-------------------------------------------------------------------------------

#Excel kitapligini yukluyoruz
import openpyxl
import sys

# bir satirdaki bilgiyi silen fonksiyon
def remove(sheet, row):
	sheet.delete_rows(row,1)


def main():
    pass

if __name__ == '__main__':
    main()

#Excel dosyası yeri ve adi tanımlanır
ExcelDosyasi = r"C:\Database\rehber.xlsx"
ExcelSayfasi = r"Sheet1"
SilinecekSatir = 2

# workbook ve worksheet olusturyoruz
workbook  = openpyxl.load_workbook(ExcelDosyasi)
worksheet = workbook[ExcelSayfasi]

print(str(SilinecekSatir)+" satir silinecek")
print("Silinme işlemi öncesi kayıt sayısı :", worksheet.max_row)

# iterate the sheet object
remove(worksheet,SilinecekSatir)

print("Silinme işlemi sonrası kayıt sayısı :",worksheet.max_row)

# Excel dosyası kayıt ediliyor
workbook.save(ExcelDosyasi)

# Excel dosyası kapatiliyor
workbook.close

# Programdan cikiliyor
sys.exit
